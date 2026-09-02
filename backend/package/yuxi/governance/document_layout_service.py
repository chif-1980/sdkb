"""统一的资料版式读取、页面预览和可编辑块模型。

版式数据只描述审核用的“视图草稿”，不会改写飞书原文件。Office 文档的视觉还原
使用 LibreOffice 转换为 PDF，PDF 与图片直接按页渲染，Excel 额外保留工作表和单元格
坐标，便于审核人员定位、修改抽取内容。
"""

from __future__ import annotations

import io
import re
import struct
from collections.abc import Sequence
from pathlib import PurePosixPath
from typing import Any

import fitz

from yuxi.governance.presentation_layout_service import extract_pptx_layout
from yuxi.services.file_preview import (
    convert_office_to_pdf,
    detect_media_type,
)

_OFFICE_EXTENSIONS = frozenset({".docx", ".xlsx", ".pptx"})
_IMAGE_EXTENSIONS = frozenset({".png", ".jpg", ".jpeg", ".gif", ".bmp", ".webp", ".svg"})
_MARKDOWN_EXTENSIONS = frozenset({".md", ".markdown"})
_LAYOUT_EXTENSIONS = frozenset(
    {".docx", ".xlsx", ".pptx", ".pdf", *_IMAGE_EXTENSIONS, *_MARKDOWN_EXTENSIONS}
)

# A source image can be kept at its original fidelity for traceability, but a
# review preview must not decode an unbounded raster in the API process or the
# browser.  These limits keep the generated preview comfortably below Pillow's
# decompression-bomb threshold while retaining enough detail for text review.
_MAX_LAYOUT_IMAGE_PIXELS = 24_000_000
_MAX_LAYOUT_IMAGE_DIMENSION = 8192
_MAX_RENDER_PAGE_PIXELS = 20_000_000
_MAX_RENDER_PAGE_DIMENSION = 6400


def _clean_text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def _normalized_text(value: Any) -> str:
    return re.sub(r"[^0-9a-z\u3400-\u9fff]+", "", _clean_text(value).lower())


def _segment_value(segment: Any, name: str, default: Any = None) -> Any:
    if isinstance(segment, dict):
        return segment.get(name, default)
    return getattr(segment, name, default)


def _matching_segment_ids(block_text: str, segments: Sequence[Any]) -> list[str]:
    normalized = _normalized_text(block_text)
    if len(normalized) < 2:
        return []
    matches: list[tuple[float, str]] = []
    for segment in segments:
        segment_id = str(_segment_value(segment, "segment_id", ""))
        content = _normalized_text(_segment_value(segment, "content", ""))
        if not segment_id or len(content) < 2:
            continue
        if normalized in content or content in normalized:
            matches.append((min(len(normalized), len(content)) / max(len(normalized), len(content)), segment_id))
    matches.sort(reverse=True)
    return [segment_id for _score, segment_id in matches[:3]]


def _percent(value: float, total: float) -> float:
    if total <= 0:
        return 0.0
    return round(max(0.0, min(100.0, value / total * 100)), 4)


def _block(
    block_id: str,
    content: str,
    *,
    left: float,
    top: float,
    width: float,
    height: float,
    kind: str = "text",
    locator: dict[str, Any] | None = None,
    segments: Sequence[Any] = (),
) -> dict[str, Any]:
    return {
        "block_id": block_id,
        "kind": kind,
        "content": content,
        "left": round(max(0.2, min(99.8, left)), 4),
        "top": round(max(0.2, min(99.8, top)), 4),
        "width": round(max(0.5, min(100.0 - left, width)), 4),
        "height": round(max(0.8, min(100.0 - top, height)), 4),
        "locator": dict(locator or {}),
        "source_segment_ids": _matching_segment_ids(content, segments),
        "editable": True,
    }


def _page(
    page_number: int,
    label: str,
    *,
    width: float,
    height: float,
    blocks: list[dict],
    render_mode: str = "image",
) -> dict:
    return {
        "page_number": page_number,
        "label": label,
        "width": round(width, 2),
        "height": round(height, 2),
        "aspect_ratio": round(width / height, 6) if height else None,
        "render_mode": render_mode,
        "block_count": len(blocks),
        "blocks": blocks,
    }


def _page_segments(segments: Sequence[Any], page_number: int) -> list[tuple[str, str, dict[str, Any]]]:
    result = []
    for segment in segments:
        locator = dict(_segment_value(segment, "locator_json", {}) or _segment_value(segment, "locator", {}) or {})
        if locator.get("page") not in (None, page_number):
            continue
        content = _clean_text(_segment_value(segment, "content", ""))
        segment_id = str(_segment_value(segment, "segment_id", ""))
        if content and segment_id:
            result.append((segment_id, content, locator))
    return result


def _open_image_document(content: bytes, filename: str):
    """Open an image through MuPDF metadata/rendering without Pillow decoding it."""
    filetype = PurePosixPath(filename).suffix.lower().lstrip(".") or None
    try:
        return fitz.open(stream=content, filetype=filetype)
    except Exception:
        # Some formats (notably SVG variants) are detected more reliably when
        # MuPDF is allowed to inspect the file signature itself.
        return fitz.open(stream=content)


def _header_image_dimensions(content: bytes, filename: str) -> tuple[float, float] | None:
    """Read raster dimensions from headers without decoding pixel data."""
    suffix = PurePosixPath(filename).suffix.lower()
    if suffix == ".png" and content.startswith(b"\x89PNG\r\n\x1a\n") and len(content) >= 24:
        width, height = struct.unpack(">II", content[16:24])
        return float(width), float(height)
    if suffix == ".gif" and content[:6] in {b"GIF87a", b"GIF89a"} and len(content) >= 10:
        width, height = struct.unpack("<HH", content[6:10])
        return float(width), float(height)
    if suffix == ".bmp" and content[:2] == b"BM" and len(content) >= 26:
        width, height = struct.unpack("<ii", content[18:26])
        return float(abs(width)), float(abs(height))
    if suffix in {".jpg", ".jpeg"} and content[:2] == b"\xff\xd8":
        offset = 2
        sof_markers = set(range(0xC0, 0xC4)) | set(range(0xC5, 0xC8)) | set(range(0xC9, 0xCC)) | set(range(0xCD, 0xD0))
        while offset + 9 < len(content):
            if content[offset] != 0xFF:
                offset += 1
                continue
            marker = content[offset + 1]
            offset += 2
            if marker in {0xD8, 0xD9}:
                continue
            if offset + 2 > len(content):
                break
            length = struct.unpack(">H", content[offset : offset + 2])[0]
            if length < 2 or offset + length > len(content):
                break
            if marker in sof_markers and length >= 7:
                height, width = struct.unpack(">HH", content[offset + 3 : offset + 7])
                return float(width), float(height)
            offset += length
    return None


def _image_dimensions(content: bytes, filename: str) -> tuple[float, float]:
    header_dimensions = _header_image_dimensions(content, filename)
    if header_dimensions and all(value > 0 for value in header_dimensions):
        return header_dimensions
    with _open_image_document(content, filename) as document:
        if document.page_count < 1:
            raise ValueError("图片中没有可处理页面")
        rect = document.load_page(0).rect
        return float(rect.width), float(rect.height)


def _image_preview_scale(width: float, height: float) -> float:
    pixels = max(width * height, 1.0)
    return min(
        1.0,
        _MAX_LAYOUT_IMAGE_DIMENSION / max(width, 1.0),
        _MAX_LAYOUT_IMAGE_DIMENSION / max(height, 1.0),
        (_MAX_LAYOUT_IMAGE_PIXELS / pixels) ** 0.5,
    )


def _render_image_preview(content: bytes, filename: str) -> tuple[bytes, str]:
    """Return a bounded PNG preview for an image source.

    The original bytes are never replaced in storage.  Only the review preview
    is rasterized and downscaled when its dimensions exceed the safe limits.
    """
    with _open_image_document(content, filename) as document:
        if document.page_count < 1:
            raise ValueError("图片中没有可处理页面")
        page = document.load_page(0)
        width, height = float(page.rect.width), float(page.rect.height)
        scale = _image_preview_scale(width, height)
        if scale >= 0.999:
            return content, detect_media_type(filename, content)
        pixmap = page.get_pixmap(matrix=fitz.Matrix(scale, scale), alpha=False)
        return pixmap.tobytes("png"), "image/png"


def _layout(filename: str, file_type: str, pages: list[dict], *, editable: bool = True) -> dict[str, Any]:
    return {
        "supported": True,
        "filename": filename,
        "file_type": file_type,
        "editable": editable,
        "page_count": len(pages),
        "pages": pages,
    }


def extract_pdf_layout(content: bytes, *, filename: str = "source.pdf", segments: Sequence[Any] = ()) -> dict[str, Any]:
    """Extract text blocks from each PDF page and normalize them to percentages."""
    pages: list[dict[str, Any]] = []
    with fitz.open(stream=content, filetype="pdf") as document:
        for page_index in range(document.page_count):
            page = document.load_page(page_index)
            rect = page.rect
            blocks: list[dict[str, Any]] = []
            text_dict = page.get_text("dict")
            block_number = 0
            for raw_block in text_dict.get("blocks", []):
                if raw_block.get("type") != 0:
                    continue
                lines = []
                for line in raw_block.get("lines", []):
                    lines.append(" ".join(span.get("text", "") for span in line.get("spans", [])))
                text = "\n".join(line for line in lines if _clean_text(line)).strip()
                if not text:
                    continue
                x0, y0, x1, y1 = raw_block.get("bbox", (0, 0, 0, 0))
                block_number += 1
                blocks.append(
                    _block(
                        f"page-{page_index + 1}-block-{block_number}",
                        text,
                        left=_percent(x0, rect.width),
                        top=_percent(y0, rect.height),
                        width=_percent(max(1.0, x1 - x0), rect.width),
                        height=_percent(max(1.0, y1 - y0), rect.height),
                        locator={"page": page_index + 1, "block": block_number},
                        segments=segments,
                    )
                )
            if not blocks:
                # Scanned PDFs often have no native text blocks. Reuse OCR segments so
                # the reviewer still gets clickable, editable evidence placeholders.
                fallback_segments = _page_segments(segments, page_index + 1)
                fallback_height = 100 / max(len(fallback_segments), 1)
                for block_number, (_segment_id, text, locator) in enumerate(fallback_segments, start=1):
                    blocks.append(
                        _block(
                            f"page-{page_index + 1}-ocr-{block_number}",
                            text,
                            left=2,
                            top=(block_number - 1) * fallback_height + 2,
                            width=96,
                            height=max(3, fallback_height - 3),
                            kind="ocr",
                            locator={"page": page_index + 1, "block": block_number, **locator},
                            segments=segments,
                        )
                    )
            pages.append(
                _page(
                    page_index + 1,
                    f"第 {page_index + 1} 页",
                    width=rect.width,
                    height=rect.height,
                    blocks=blocks,
                )
            )
    return _layout(filename, ".pdf", pages)


def _excel_cell_value(cell: Any) -> str:
    value = cell.value
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def extract_xlsx_layout(
    content: bytes,
    *,
    filename: str = "source.xlsx",
    segments: Sequence[Any] = (),
) -> dict[str, Any]:
    """Build a lightweight, editable sheet grid without flattening cells into paragraphs."""
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(content), data_only=False, read_only=True)
    pages: list[dict[str, Any]] = []
    try:
        for page_number, worksheet in enumerate(workbook.worksheets, start=1):
            max_row = max(worksheet.max_row or 1, 1)
            max_column = max(worksheet.max_column or 1, 1)
            blocks: list[dict[str, Any]] = []
            for row in worksheet.iter_rows():
                for cell in row:
                    value = _excel_cell_value(cell)
                    if not value.strip():
                        continue
                    left = (cell.column - 1) / max_column * 100
                    top = (cell.row - 1) / max_row * 100
                    blocks.append(
                        _block(
                            f"sheet-{page_number}-cell-{cell.coordinate}",
                            value,
                            left=left,
                            top=top,
                            width=100 / max_column,
                            height=100 / max_row,
                            kind="cell",
                            locator={
                                "sheet": worksheet.title,
                                "cell": cell.coordinate,
                                "row": cell.row,
                                "column": cell.column,
                            },
                            segments=segments,
                        )
                    )
            pages.append(
                _page(
                    page_number,
                    worksheet.title,
                    width=max_column,
                    height=max_row,
                    blocks=blocks,
                    render_mode="grid",
                )
            )
    finally:
        workbook.close()
    return _layout(filename, ".xlsx", pages)


def _image_box(locator: dict[str, Any], width: float, height: float) -> tuple[float, float, float, float] | None:
    raw_box = locator.get("bbox") or locator.get("bounding_box")
    if not isinstance(raw_box, (list, tuple)) or len(raw_box) < 4:
        return None
    x0, y0, x1, y1 = [float(value) for value in raw_box[:4]]
    if x1 <= x0 or y1 <= y0:
        return None
    # OCR engines may return either pixel coordinates or normalized percentages.
    if max(x1, y1) <= 100.0:
        return x0, y0, x1 - x0, y1 - y0
    return _percent(x0, width), _percent(y0, height), _percent(x1 - x0, width), _percent(y1 - y0, height)


def extract_image_layout(
    content: bytes,
    *,
    filename: str = "source.png",
    segments: Sequence[Any] = (),
) -> dict[str, Any]:
    # Read dimensions through MuPDF so an oversized source does not trigger
    # Pillow's decompression-bomb guard before we can create a bounded preview.
    width, height = _image_dimensions(content, filename)
    image_segments = []
    for segment in segments:
        locator = dict(_segment_value(segment, "locator_json", {}) or _segment_value(segment, "locator", {}) or {})
        page = locator.get("page")
        if page not in (None, 1):
            continue
        text = _clean_text(_segment_value(segment, "content", ""))
        if text:
            image_segments.append((locator, text, str(_segment_value(segment, "segment_id", ""))))
    blocks: list[dict[str, Any]] = []
    fallback_height = 100 / max(len(image_segments), 1)
    for index, (locator, text, _segment_id) in enumerate(image_segments, start=1):
        box = _image_box(locator, width, height)
        if box is None:
            box = (2.0, (index - 1) * fallback_height + 2.0, 96.0, max(3.0, fallback_height - 3.0))
        left, top, block_width, block_height = box
        blocks.append(
            _block(
                f"image-1-block-{index}",
                text,
                left=left,
                top=top,
                width=block_width,
                height=block_height,
                kind="ocr",
                locator={"page": 1, "block": index, **locator},
                segments=segments,
            )
        )
    page = _page(1, "图片", width=width, height=height, blocks=blocks)
    page["preview_scaled"] = _image_preview_scale(width, height) < 0.999
    page["preview_width"] = round(width * _image_preview_scale(width, height))
    page["preview_height"] = round(height * _image_preview_scale(width, height))
    return _layout(filename, PurePosixPath(filename).suffix.lower(), [page])


def _pptx_layout(content: bytes, *, filename: str, segments: Sequence[Any]) -> dict[str, Any]:
    presentation = extract_pptx_layout(content, segments=segments).as_dict()
    pages = [
        _page(
            slide["slide_number"],
            f"第 {slide['slide_number']} 页",
            width=presentation["slide_width"],
            height=presentation["slide_height"],
            blocks=[
                {
                    **fragment,
                    "block_id": fragment["fragment_id"],
                    "kind": "text",
                    "locator": {"slide": slide["slide_number"], "shape": fragment.get("shape_number")},
                    "editable": True,
                }
                for fragment in slide["fragments"]
            ],
        )
        for slide in presentation["slides"]
    ]
    return _layout(filename, ".pptx", pages)


def _split_markdown_blocks(markdown: str) -> list[str]:
    blocks: list[str] = []
    current: list[str] = []
    fence_marker = ""

    def flush() -> None:
        content = "\n".join(current).strip()
        if content:
            blocks.append(content)
        current.clear()

    for line in markdown.splitlines():
        stripped = line.lstrip()
        marker = stripped[:3] if stripped.startswith(("```", "~~~")) else ""
        if marker:
            if not fence_marker:
                fence_marker = marker
            elif marker == fence_marker:
                fence_marker = ""
            current.append(line)
            continue
        if not stripped and not fence_marker:
            flush()
            continue
        current.append(line)
    flush()
    return blocks


def extract_markdown_layout(
    content: bytes,
    *,
    filename: str = "source.md",
    segments: Sequence[Any] = (),
) -> dict[str, Any]:
    """Build a semantic review layout for Markdown, which has no fixed page geometry."""
    markdown = content.decode("utf-8-sig", errors="replace")
    source_blocks = _split_markdown_blocks(markdown)
    block_height = 100 / max(len(source_blocks), 1)
    blocks = [
        _block(
            f"markdown-block-{index}",
            block_content,
            left=2,
            top=(index - 1) * block_height + 1,
            width=96,
            height=max(1, block_height - 2),
            kind="markdown",
            locator={"section": index},
            segments=segments,
        )
        for index, block_content in enumerate(source_blocks, start=1)
    ]
    page = _page(
        1,
        "Markdown 正文",
        width=720,
        height=max(len(markdown.splitlines()), 1),
        blocks=blocks,
        render_mode="markdown",
    )
    return _layout(filename, PurePosixPath(filename).suffix.lower(), [page], editable=False)


async def build_document_layout(
    filename: str,
    content: bytes,
    *,
    segments: Sequence[Any] = (),
) -> dict[str, Any]:
    """Parse one supported source into the common review layout model."""
    suffix = PurePosixPath(filename).suffix.lower()
    if suffix == ".pptx":
        return _pptx_layout(content, filename=filename, segments=segments)
    if suffix == ".pdf":
        return extract_pdf_layout(content, filename=filename, segments=segments)
    if suffix == ".xlsx":
        return extract_xlsx_layout(content, filename=filename, segments=segments)
    if suffix in _IMAGE_EXTENSIONS:
        return extract_image_layout(content, filename=filename, segments=segments)
    if suffix == ".docx":
        pdf_content = await convert_office_to_pdf(filename, content)
        layout = extract_pdf_layout(pdf_content, filename=filename, segments=segments)
        layout["file_type"] = ".docx"
        return layout
    if suffix in _MARKDOWN_EXTENSIONS:
        return extract_markdown_layout(content, filename=filename, segments=segments)
    return {
        "supported": False,
        "filename": filename,
        "file_type": suffix,
        "editable": False,
        "page_count": 0,
        "pages": [],
        "message": "当前文件格式暂不支持版式还原",
    }


async def render_document_page(
    filename: str,
    content: bytes,
    *,
    page_number: int,
    pdf_content: bytes | None = None,
    density: int = 1,
) -> tuple[bytes, str]:
    """Render a page as PNG (or return the original image bytes)."""
    suffix = PurePosixPath(filename).suffix.lower()
    if suffix in _IMAGE_EXTENSIONS:
        if page_number != 1:
            raise IndexError("图片只有一页")
        return _render_image_preview(content, filename)
    if suffix in {".docx", ".pptx"}:
        if pdf_content is None:
            pdf_content = await convert_office_to_pdf(filename, content)
    elif suffix == ".xlsx":
        raise ValueError("Excel 使用工作表网格预览，无需生成页面图片")
    else:
        pdf_content = content
    with fitz.open(stream=pdf_content, filetype="pdf") as document:
        if page_number < 1 or page_number > document.page_count:
            raise IndexError("页码超出范围")
        page = document.load_page(page_number - 1)
        density = max(1, min(3, int(density)))
        width = max(float(page.rect.width), 1.0)
        height = max(float(page.rect.height), 1.0)
        target_scale = max(1.0, (1600 * density) / width)
        scale = min(
            target_scale,
            _MAX_RENDER_PAGE_DIMENSION / width,
            _MAX_RENDER_PAGE_DIMENSION / height,
            (_MAX_RENDER_PAGE_PIXELS / (width * height)) ** 0.5,
        )
        pixmap = page.get_pixmap(matrix=fitz.Matrix(scale, scale), alpha=False)
        return pixmap.tobytes("jpeg", jpg_quality=94 if density > 1 else 90), "image/jpeg"


def supported_layout_suffix(suffix: str) -> bool:
    return suffix.lower() in _LAYOUT_EXTENSIONS
